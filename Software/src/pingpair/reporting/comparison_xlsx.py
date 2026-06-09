"""XLSX writer for the Analysis-tab comparison report (#12).

One workbook, three sheets:

* **Summary** — run-level rows (label / started / duration / cases ok)
  plus the filter snapshot.
* **Stats** — one row per run, columns by metric (min / avg / median /
  max / stdev). Median + stdev are present here in addition to the
  three-column display on docx/txt so the analyst can dig deeper.
* **Diff** — only when len(runs) == 2; one row per case with A / B / Δ
  triples per metric.

Embedded PNGs are not added to xlsx — openpyxl supports it but the
typical use is "compute and re-pivot in Excel," for which raw cells
are more useful than embedded images.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..analysis import METRICS
from ..analysis.comparison import ComparisonReport
from .run_report import fmt_duration


_HEADER_FILL = PatternFill(
    start_color="FF2A3F66", end_color="FF2A3F66", fill_type="solid"
)
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        max_w = 8
        for cell in ws[letter]:
            value = cell.value
            if value is not None:
                max_w = max(max_w, len(str(value)) + 2)
        ws.column_dimensions[letter].width = min(max_w, 40)


def write_comparison_xlsx(report: ComparisonReport, dest: Path) -> None:
    """Write ``report`` to ``dest`` as an .xlsx file."""
    wb = Workbook()

    # ---- Summary sheet ----
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = report.title
    summary["A1"].font = Font(bold=True, size=14)
    summary["A2"] = (
        f"Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
    )
    summary["A3"] = f"Runs compared: {report.run_count}"
    if not report.filter_description.is_default:
        summary["A5"] = "Filter applied:"
        summary["A5"].font = Font(bold=True)
        for i, line in enumerate(report.filter_description.lines(), start=6):
            summary[f"A{i}"] = f"  · {line}"
    if report.notes:
        offset = summary.max_row + 2
        summary.cell(row=offset, column=1, value="Notes:").font = Font(bold=True)
        for i, line in enumerate(report.notes.splitlines(), start=offset + 1):
            summary.cell(row=i, column=1, value=line)

    # Runs included table at the bottom of Summary.
    runs_top = summary.max_row + 2
    headers = ["Run", "Started", "Duration", "Type", "Cases ok"]
    for c, h in enumerate(headers, start=1):
        summary.cell(row=runs_top, column=c, value=h)
    _style_header_row(summary, runs_top, len(headers))
    for r_idx, run in enumerate(report.runs, start=runs_top + 1):
        when = (
            run.started_at.strftime("%Y-%m-%d %H:%M")
            if run.started_at is not None
            else "?"
        )
        summary.cell(row=r_idx, column=1, value=run.display_label)
        summary.cell(row=r_idx, column=2, value=when)
        summary.cell(row=r_idx, column=3, value=fmt_duration(run.duration_s))
        summary.cell(
            row=r_idx, column=4,
            value="multi" if run.is_multi_segment else "single",
        )
        summary.cell(
            row=r_idx, column=5,
            value=f"{run.cases_ok}/{run.cases_total}",
        )
    _autosize(summary, len(headers))

    # ---- Stats sheet ----
    stats_ws = wb.create_sheet("Stats")
    headers = ["Run"]
    for metric in METRICS:
        short = metric.display.split()[0]
        headers.extend([
            f"{short} min ({metric.unit})",
            f"{short} avg ({metric.unit})",
            f"{short} median ({metric.unit})",
            f"{short} max ({metric.unit})",
            f"{short} stdev ({metric.unit})",
            f"{short} samples",
        ])
    for c, h in enumerate(headers, start=1):
        stats_ws.cell(row=1, column=c, value=h)
    _style_header_row(stats_ws, 1, len(headers))
    for r_idx, (run, rs) in enumerate(
        zip(report.runs, report.per_run_stats, strict=False), start=2
    ):
        stats_ws.cell(row=r_idx, column=1, value=run.display_label)
        col = 2
        for metric in METRICS:
            ms = rs.by_metric[metric.code]
            for v in (ms.min, ms.avg, ms.median, ms.max, ms.stdev):
                stats_ws.cell(row=r_idx, column=col, value=v)
                col += 1
            stats_ws.cell(row=r_idx, column=col, value=ms.samples)
            col += 1
    _autosize(stats_ws, len(headers))

    # ---- Diff sheet (only when len(runs) == 2) ----
    if report.has_diff_section:
        diff_ws = wb.create_sheet("Diff")
        a = report.runs[1]
        b = report.runs[0]
        diff_ws["A1"] = (
            f"A (older): {a.display_label}    B (newer): {b.display_label}"
        )
        diff_ws["A1"].font = Font(bold=True)
        headers = ["Case", "Payload (B)", "BW (Mbps)"]
        for metric in METRICS:
            short = metric.display.split()[0]
            headers.extend([
                f"A {short}", f"B {short}", f"Δ {short}",
            ])
        for c, h in enumerate(headers, start=1):
            diff_ws.cell(row=3, column=c, value=h)
        _style_header_row(diff_ws, 3, len(headers))
        for r_idx, row in enumerate(report.per_case_diff_rows, start=4):
            diff_ws.cell(row=r_idx, column=1, value=row.case_idx)
            diff_ws.cell(row=r_idx, column=2, value=row.payload_bytes)
            diff_ws.cell(
                row=r_idx, column=3, value=row.bandwidth_mbps_pushed,
            )
            col = 4
            for metric in METRICS:
                diff_ws.cell(
                    row=r_idx, column=col,
                    value=row.a_value[metric.code],
                )
                diff_ws.cell(
                    row=r_idx, column=col + 1,
                    value=row.b_value[metric.code],
                )
                diff_ws.cell(
                    row=r_idx, column=col + 2,
                    value=row.delta[metric.code],
                )
                col += 3
        _autosize(diff_ws, len(headers))

    wb.save(str(dest))
