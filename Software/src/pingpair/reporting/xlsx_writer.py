"""Render a :class:`RunReport` as an Excel workbook.

Three sheets:

* **Summary** — the canonical 8-column table from Table-1.PNG, one row
  per test case. Suitable for charting and quick comparison.
* **Detail** — every column from the docx detail section flattened into
  one row per case (status, return codes, exact CLI strings used,
  per-case error if any).
* **Run info** — the run-level metadata: timestamps, IPs, software
  versions, protocol, cases-ok counters, and the full AppConfig snapshot
  used for the run.

We use openpyxl rather than pandas to avoid a heavy dep just to write
two tables.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .run_report import (
    CaseMetrics,
    MultiSegmentRunReport,
    RunReport,
    SEGMENT_SUMMARY_HEADERS,
    cross_segment_comparison,
    fmt_duration,
    metadata_rows,
    segment_summary_rows,
)


# Header style for all three sheets.
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF305496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


_HEADERS_SUMMARY = (
    "#",
    "Payload (B)",
    "Bandwidth Pushed (Mbps)",
    "Throughput Received (Mbps)",
    "Jitter (ms)",
    "Packet Loss (%)",
    "Min Latency (ms)",
    "Avg Latency (ms)",
    "Max Latency (ms)",
)

_HEADERS_DETAIL = (
    "#",
    "Payload (B)",
    "BW Pushed (Mbps)",
    "Duration (s)",
    "Status",
    "Error",
    "Throughput (Mbps)",
    "Jitter (ms)",
    "Loss (%)",
    "Min (ms)",
    "Avg (ms)",
    "Max (ms)",
    "iperf3 client rc",
    "iperf3 server rc",
    "fping rc",
    "iperf3 client cmd",
    "iperf3 server cmd",
    "fping cmd",
)


def _none(v: object) -> object:
    """Map None → empty string so Excel doesn't render them as #N/A."""
    return "" if v is None else v


def _write_headers(ws: Worksheet, headers: tuple[str, ...]) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    """Set column widths from the longest cell value, capped at max_width."""
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        longest = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------


def _write_summary(ws: Worksheet, report: RunReport) -> None:
    ws.title = "Summary"
    _write_headers(ws, _HEADERS_SUMMARY)
    for r, c in enumerate(report.cases, start=2):
        ws.cell(row=r, column=1, value=c.case_idx)
        ws.cell(row=r, column=2, value=c.payload_bytes)
        ws.cell(row=r, column=3, value=c.bandwidth_mbps_pushed)
        ws.cell(row=r, column=4, value=_none(c.throughput_mbps_received))
        ws.cell(row=r, column=5, value=_none(c.jitter_ms))
        ws.cell(row=r, column=6, value=_none(c.packet_loss_pct))
        ws.cell(row=r, column=7, value=_none(c.min_latency_ms))
        ws.cell(row=r, column=8, value=_none(c.avg_latency_ms))
        ws.cell(row=r, column=9, value=_none(c.max_latency_ms))
    # Round numeric columns to a sensible precision.
    for col, fmt in ((4, "0.00"), (5, "0.000"), (6, "0.000"),
                     (7, "0.00"), (8, "0.00"), (9, "0.00")):
        letter = get_column_letter(col)
        for r in range(2, len(report.cases) + 2):
            ws[f"{letter}{r}"].number_format = fmt
    _autosize(ws)


def _write_detail(ws: Worksheet, report: RunReport) -> None:
    ws.title = "Detail"
    _write_headers(ws, _HEADERS_DETAIL)
    for r, c in enumerate(report.cases, start=2):
        ws.cell(row=r, column=1, value=c.case_idx)
        ws.cell(row=r, column=2, value=c.payload_bytes)
        ws.cell(row=r, column=3, value=c.bandwidth_mbps_pushed)
        ws.cell(row=r, column=4, value=c.duration_s)
        ws.cell(row=r, column=5, value=c.status)
        ws.cell(row=r, column=6, value=c.error or "")
        ws.cell(row=r, column=7, value=_none(c.throughput_mbps_received))
        ws.cell(row=r, column=8, value=_none(c.jitter_ms))
        ws.cell(row=r, column=9, value=_none(c.packet_loss_pct))
        ws.cell(row=r, column=10, value=_none(c.min_latency_ms))
        ws.cell(row=r, column=11, value=_none(c.avg_latency_ms))
        ws.cell(row=r, column=12, value=_none(c.max_latency_ms))
        ws.cell(row=r, column=13, value=_none(c.iperf3_client_rc))
        ws.cell(row=r, column=14, value=_none(c.iperf3_server_rc))
        ws.cell(row=r, column=15, value=_none(c.fping_rc))
        ws.cell(row=r, column=16, value=c.iperf3_client_cmd)
        ws.cell(row=r, column=17, value=c.iperf3_server_cmd)
        ws.cell(row=r, column=18, value=c.fping_cmd)
    _autosize(ws, max_width=80)


def _write_run_info(ws: Worksheet, report: RunReport) -> None:
    ws.title = "Run info"
    _write_headers(ws, ("Field", "Value"))

    rows: list[tuple[str, object]] = [
        ("Run ID", report.run_id),
        ("Started", report.started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Finished", report.ended_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Duration", fmt_duration(report.duration_s)),
        ("Duration (s, raw)", round(report.duration_s, 2)),
        ("Server IP", report.server_ip),
        ("Client IP", report.client_ip),
        ("Protocol", report.protocol.upper()),
        ("Cases ok", report.cases_ok),
        ("Cases total", report.cases_total),
        ("fping version", report.fping_version),
        ("iperf3 version", report.iperf3_version),
        ("PingPair version", report.app_version),
    ]
    # Test-record metadata, only the populated fields.
    rows.extend(metadata_rows(report))
    rows.append((
        "AppConfig snapshot",
        json.dumps(report.cfg_snapshot, indent=2, default=str),
    ))
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    _autosize(ws, max_width=120)


# ---------------------------------------------------------------------------
# Top level
# ---------------------------------------------------------------------------


def write_xlsx(
    report: RunReport,
    dest: Path,
    *,
    include_appendix: bool = False,
) -> None:
    wb = Workbook()
    summary = wb.active
    _write_summary(summary, report)
    _write_detail(wb.create_sheet("Detail"), report)
    _write_run_info(wb.create_sheet("Run info"), report)

    if include_appendix:
        from .appendix import append_to_xlsx
        append_to_xlsx(wb, report, is_multi=False)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))


# ---------------------------------------------------------------------------
# Group C-1: multi-segment writer
# ---------------------------------------------------------------------------


def _sanitize_sheet_name(name: str) -> str:
    """Strip / shorten a string so openpyxl accepts it as a sheet name.

    Excel forbids these chars: [ ] : * ? / \\ and caps the name at 31
    chars. We replace forbidden chars with '-' and truncate. Mirrors
    what Excel itself does when you paste a long sheet name.
    """
    bad = set("[]:*?/\\")
    cleaned = "".join("-" if c in bad else c for c in name)
    return cleaned[:31] or "Segment"


def _write_multi_summary(ws: Worksheet, report: MultiSegmentRunReport) -> None:
    """Top-level Summary: segments table + run-level info."""
    ws.title = "Summary"

    # Section 1: segments summary table.
    _write_headers(ws, SEGMENT_SUMMARY_HEADERS)
    for r, row_values in enumerate(segment_summary_rows(report), start=2):
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=r, column=col, value=value)

    # A few blank rows, then the run-level info — keeps the summary
    # scannable without forcing the operator to open another sheet.
    info_start = len(report.segments) + 4
    ws.cell(row=info_start - 1, column=1, value="Run information").font = Font(bold=True)
    info_rows: list[tuple[str, object]] = [
        ("Run ID", report.run_id),
        ("Started", report.started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Finished", report.ended_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Total duration", fmt_duration(report.duration_s)),
        ("Server IP", report.server_ip),
        ("Client IP", report.client_ip),
        ("Protocol", report.protocol.upper()),
        ("Segments ok", f"{report.segments_ok} / {report.segments_total}"),
        ("Cases ok", f"{report.total_cases_ok} / {report.total_cases}"),
        ("PingPair version", report.app_version),
    ]
    info_rows.extend(metadata_rows(report))
    for offset, (k, v) in enumerate(info_rows):
        ws.cell(row=info_start + offset, column=1, value=k)
        ws.cell(row=info_start + offset, column=2, value=v)

    _autosize(ws)


def _write_segment_detail(
    ws: Worksheet, segment_idx: int, label: str, cases: list[CaseMetrics]
) -> None:
    """One sheet per segment — same shape as the single-sweep Detail sheet."""
    ws.title = _sanitize_sheet_name(f"Seg {segment_idx} - {label}")
    _write_headers(ws, _HEADERS_DETAIL)
    for r, c in enumerate(cases, start=2):
        ws.cell(row=r, column=1, value=c.case_idx)
        ws.cell(row=r, column=2, value=c.payload_bytes)
        ws.cell(row=r, column=3, value=c.bandwidth_mbps_pushed)
        ws.cell(row=r, column=4, value=c.duration_s)
        ws.cell(row=r, column=5, value=c.status)
        ws.cell(row=r, column=6, value=c.error or "")
        ws.cell(row=r, column=7, value=_none(c.throughput_mbps_received))
        ws.cell(row=r, column=8, value=_none(c.jitter_ms))
        ws.cell(row=r, column=9, value=_none(c.packet_loss_pct))
        ws.cell(row=r, column=10, value=_none(c.min_latency_ms))
        ws.cell(row=r, column=11, value=_none(c.avg_latency_ms))
        ws.cell(row=r, column=12, value=_none(c.max_latency_ms))
        ws.cell(row=r, column=13, value=_none(c.iperf3_client_rc))
        ws.cell(row=r, column=14, value=_none(c.iperf3_server_rc))
        ws.cell(row=r, column=15, value=_none(c.fping_rc))
        ws.cell(row=r, column=16, value=c.iperf3_client_cmd)
        ws.cell(row=r, column=17, value=c.iperf3_server_cmd)
        ws.cell(row=r, column=18, value=c.fping_cmd)
    _autosize(ws, max_width=80)


def _write_comparison_sheet(
    ws: Worksheet,
    report: MultiSegmentRunReport,
    *,
    metric_id: str,
    metric_label: str,
) -> None:
    """One sheet per metric: rows = case, columns = segments."""
    ws.title = _sanitize_sheet_name(metric_label)

    # First row: blank A1, then full metric label (banner-style).
    ws.cell(row=1, column=1, value=metric_label).font = Font(bold=True, size=12)
    ws.row_dimensions[1].height = 22

    # ``as_values=True`` yields native numeric cells (int case#/payload/bw,
    # float|None metric cells) — so the sheet and its charts carry the
    # full-precision sidecar values instead of numbers round-tripped
    # through a 2/3-decimal display string.
    headers, rows = cross_segment_comparison(
        report, metric=metric_id, as_values=True
    )
    # Write headers on row 2 (banner row 1).
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "D3"  # freeze case# / payload / bw + header row

    for r, row_values in enumerate(rows, start=3):
        for col, value in enumerate(row_values, start=1):
            # A missing case is None → an empty cell, which charts skip.
            ws.cell(row=r, column=col, value=value)
    _autosize(ws)


def _write_multi_run_info(ws: Worksheet, report: MultiSegmentRunReport) -> None:
    ws.title = "Run info"
    _write_headers(ws, ("Field", "Value"))

    rows: list[tuple[str, object]] = [
        ("Run ID", report.run_id),
        ("Started", report.started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Finished", report.ended_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("Total duration", fmt_duration(report.duration_s)),
        ("Total duration (s, raw)", round(report.duration_s, 2)),
        ("Server IP", report.server_ip),
        ("Client IP", report.client_ip),
        ("Protocol", report.protocol.upper()),
        ("Segments ok", report.segments_ok),
        ("Segments total", report.segments_total),
        ("Total cases ok", report.total_cases_ok),
        ("Total cases", report.total_cases),
        ("Selected case indexes", ",".join(str(i) for i in report.selected_case_indexes) or "(all 20)"),
        ("fping version", report.fping_version),
        ("iperf3 version", report.iperf3_version),
        ("PingPair version", report.app_version),
    ]
    rows.extend(metadata_rows(report))
    rows.append((
        "AppConfig snapshot",
        json.dumps(report.cfg_snapshot, indent=2, default=str),
    ))
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    _autosize(ws, max_width=120)


def write_multi_xlsx(
    report: MultiSegmentRunReport,
    dest: Path,
    *,
    include_appendix: bool = False,
) -> None:
    """Write a multi-segment .xlsx workbook to ``dest``.

    Sheets, in order:
      Summary               – segments table + run-level info banner
      Throughput (Mbps)     – pivoted comparison, rows = case, cols = segment
      Avg Latency (ms)      – same shape
      Packet Loss (%)       – same shape
      Seg N - <label>       – one sheet per segment, full per-case detail
      Run info              – run-level metadata + AppConfig snapshot
    """
    wb = Workbook()

    summary = wb.active
    _write_multi_summary(summary, report)

    # Comparison sheets next so they're easy to find when the user
    # opens the workbook from the file manager.
    for metric_id, metric_label in (
        ("throughput", "Throughput (Mbps)"),
        ("avg_latency", "Avg Latency (ms)"),
        ("loss", "Packet Loss (%)"),
    ):
        _write_comparison_sheet(
            wb.create_sheet(metric_label),
            report,
            metric_id=metric_id,
            metric_label=metric_label,
        )


    # Per-segment Detail sheets.
    for seg in report.segments:
        _write_segment_detail(
            wb.create_sheet(),
            segment_idx=seg.segment_idx,
            label=seg.label,
            cases=seg.cases,
        )

    _write_multi_run_info(wb.create_sheet("Run info"), report)

    if include_appendix:
        from .appendix import append_to_xlsx
        append_to_xlsx(wb, report, is_multi=True)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(dest))
